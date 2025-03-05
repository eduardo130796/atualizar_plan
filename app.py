import pandas as pd
import openpyxl
from datetime import datetime
import io
import streamlit as st

# Função principal para processamento dos arquivos
@st.cache_data
def processar_planilhas(planilha_base, planilha_atualizacao):
    wb = openpyxl.load_workbook(planilha_base)
    ws = wb.active  # Aba principal
      # Nome original da planilha base

    if "Log de Alterações" not in wb.sheetnames:
        ws_log = wb.create_sheet("Log de Alterações")
        ws_log.append(["Tipo", "Nota de Empenho", "Campo", "Valor Antigo", "Valor Novo", "Data da Alteração"])

    df_atualizacao = pd.read_excel(planilha_atualizacao, skiprows=2)
    
    mapa_valor_empenhado = {
        str(row["Número da Nota de Empenho"]).strip()[-8:]: row["Saldo - R$ (Item Informação)"]
        for _, row in df_atualizacao.iterrows()
    }

    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[4]
        valor_empenhado_cell = row[5]
        nota_empenho = str(nota_empenho_cell.value).strip()

        if nota_empenho in mapa_valor_empenhado:
            novo_valor = mapa_valor_empenhado[nota_empenho]
            if str(valor_empenhado_cell.value).strip() != str(novo_valor).strip():
                ws_log.append(["Empenho", nota_empenho, "Valor Empenhado", valor_empenhado_cell.value, novo_valor, 
                               pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
                valor_empenhado_cell.value = novo_valor

    df_atualizacao = df_atualizacao[~df_atualizacao.apply(lambda row: row.astype(str).str.contains('Total').any(), axis=1)]
    df_atualizacao.ffill(inplace=True)

    meses = {"jan": 12, "fev": 13, "mar": 14, "abr": 15, "mai": 16, "jun": 17,
             "jul": 18, "ago": 19, "set": 20, "out": 21, "nov": 22, "dez": 23}
    meses_ingles_para_portugues = {'jan': 'jan', 'feb': 'fev', 'mar': 'mar', 'apr': 'abr', 'may': 'mai',
                                   'jun': 'jun', 'jul': 'jul', 'aug': 'ago', 'sep': 'set', 'oct': 'out',
                                   'nov': 'nov', 'dec': 'dez'}

    pagamentos_por_nota = {}
    for _, row in df_atualizacao.iterrows():
        nota_empenho = str(row["Número da Nota de Empenho"]).strip()[-8:]
        data_pagamento = row["Métrica"]
        valor_pago = row["Unnamed: 13"]

        if nota_empenho not in pagamentos_por_nota:
            pagamentos_por_nota[nota_empenho] = {mes: [] for mes in meses}

        data_pagamento = pd.to_datetime(data_pagamento, errors='coerce', dayfirst=True)
        if pd.notna(data_pagamento):
            mes_pagamento = data_pagamento.strftime('%b').lower()
            mes_pagamento_portugues = meses_ingles_para_portugues.get(mes_pagamento)
            if mes_pagamento_portugues in meses:
                pagamentos_por_nota[nota_empenho][mes_pagamento_portugues].append(float(valor_pago))

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
        nota_empenho_cell = row[4]
        status_cell = row[10]
        nota_empenho = str(nota_empenho_cell.value).strip()

        if nota_empenho in pagamentos_por_nota:
            for mes, coluna_mes in meses.items():
                valores_novos = pagamentos_por_nota[nota_empenho].get(mes, [])

                if valores_novos:
                    valor_pago_cell = row[coluna_mes]

                    if isinstance(valores_novos, (int, float)):
                        valores_novos = [valores_novos]

                    status_texto = str(status_cell.value).strip().lower() if status_cell.value else ""

                    status_aceitos = [
                        "não pediu, mas pode solicitar.",
                        "solicitado - em análise",
                        "Não solicitou, mas pode pedir"
                    ]

                    if any(status_texto.startswith(opcao.lower()) for opcao in status_aceitos):
                        soma_valores = "+".join(str(v).replace(",", ".") for v in valores_novos)
                        nova_formula = f"=({soma_valores})+(({soma_valores})*AE6)"
                    else:
                        soma_valores = "+".join(str(v).replace(",", ".") for v in valores_novos)
                        nova_formula = f"={soma_valores}" if len(valores_novos) > 1 else f"={valores_novos[0]}"

                    if str(valor_pago_cell.value).strip() != nova_formula:
                        ws_log.append(["Pagamento", nota_empenho, f"Pagamento {mes}", valor_pago_cell.value, nova_formula, 
                                       pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')])
                        valor_pago_cell.value = nova_formula

    data_hora_atualizacao = datetime.now().strftime("Última atualização: %d/%m/%Y às %H:%M")
    ws["A1"] = data_hora_atualizacao

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return buffer

# Interface Streamlit
st.set_page_config(page_title="Processamento de Planilhas", page_icon="📊", layout="centered")
st.title("🚀 Processamento de Planilhas de Orçamento")
st.markdown("""
    **Bem-vindo à ferramenta interativa de processamento de planilhas!**
    
    Faça o upload da **planilha com o objeto  desejado** e da **planilha com as Notas de Empenho** para gerar uma planilha final com as informações processadas.
""")

st.markdown("""
    ### Instruções para uso:
    1. **Selecione** os dois arquivos no campo correspondente.
    2. O sistema irá processar automaticamente a planilha e atualizar os dados de acordo com as informações fornecidas.
    3. Após o processamento, você pode **baixar a planilha finalizada**.

    """)

col1, col2 = st.columns([2, 2])
with col1:
    uploaded_file_base = st.file_uploader("📂 Selecione a planilha com o Objeto", type=["xlsx"])
with col2:
    uploaded_file_atualizacao = st.file_uploader("📊 Selecione a planilha de Notas de Empenho", type=["xlsx"])

if uploaded_file_base and uploaded_file_atualizacao:
    st.info("💡 Processando os dados... Isso pode levar alguns minutos.")

    with st.spinner("🔄 Processando, por favor aguarde..."):
        buffer_final = processar_planilhas(uploaded_file_base, uploaded_file_atualizacao)

    st.success("✅ Planilha processada com sucesso! Pronto para o download.")
    nome_arquivo_saida = uploaded_file_base.name.replace(".xlsx", "_atualizada.xlsx")
    # Colocando um botão de download
    st.download_button(
        label="📥 Baixar Planilha Final Atualizada",
        data=buffer_final,
        file_name=nome_arquivo_saida,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    st.markdown("""
       
    ### Dúvidas?
    Se você tiver problemas com a ferramenta, entre em contato com o suporte.
    """)

else:
    st.warning("⚠️ Para processar, por favor, faça o upload de ambos os arquivos acima.")


# Rodapé fixo com largura total
rodape = """
    <style>
        .footer {
            position: fixed;
            left: 0;
            bottom: 0;
            width: 100%;
            background-color: #f8f9fa;
            text-align: center;
            padding: 10px;
            font-size: 14px;
            color: #6c757d;
            border-top: 1px solid #dee2e6;
            z-index: 100;
        }
    </style>
    <div class="footer">
        Desenvolvido por <strong>Eduardo Júnior</strong> | 2025
    </div>
"""

# Exibir o rodapé na interface
st.markdown(rodape, unsafe_allow_html=True)
